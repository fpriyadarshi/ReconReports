'''
Created on Apr 17, 2019

@author: fagoon_priyadarshi
'''
from selenium.common.exceptions import ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.support.select import Select
from chromedriver_py import binary_path
import zipfile
import xlsxwriter
from collections import Counter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import yaml
import time
import pytest
import pytest_html
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
import calendar
from datetime import date
from datetime import datetime
from tabulate import tabulate
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib, ssl
import base64
import mimetypes
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import os, shutil, fnmatch, csv
import logging
import sys, pdb
from openpyxl.styles import PatternFill
import win32serviceutil
from yattag import Doc
from os import path
import pdb
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver

cwd = os.path.dirname(os.path.dirname(__file__))

downloadPath = cwd + "\\" + "reports\\"
reconReportDownloadPath = cwd + "\\" + "reports_for_recon\\"

if not os.path.exists(downloadPath):
    try:
        os.mkdir(downloadPath)
    except OSError:
        print("Creation of the directory %s failed" % downloadPath)
    else:
        print("Successfully created the directory %s " % downloadPath)

if not os.path.exists(reconReportDownloadPath):
    try:
        os.mkdir(reconReportDownloadPath)
    except OSError:
        print("Creation of the recon files download directory %s failed" % reconReportDownloadPath)
    else:
        print("Successfully created the recon files download directory %s " % reconReportDownloadPath)        


# currentDateTime = datetime.now()
# logger_file = cwd + "\\" + "logs\\" + f"RR_LOG_{currentDateTime}.txt"
# logging.basicConfig(filename=logger_file, filemode='w', level=logging.INFO,
#                     format='%(message)s')
# logging.info("\n======================================================================================================")
# logging.info("\t\t\t NEW LOG:===>  " + str(currentDateTime))
# logging.info("======================================================================================================")
# yellowFill = PatternFill(fill_type='solid', start_color="FFD700", end_color="FF5733")

class Test_URL():
    recordCountMap = {}
    driver = None
    driver_wait = None
    logger = None
    today = datetime.today()
    oppDateToday = today.strftime("%Y%m%d%H%M%S")            
    loggerFileName = os.path.dirname(os.path.dirname(__file__)) + "\\logs\\" + f"RR_LOG_{oppDateToday}.txt"

    def __init__(self):
        try:
            # Initiate Logger
            

            self.logger = logging.getLogger(os.getenv('ORG_TYPE'))
            
            if not self.logger.handlers:    
                self.logger.setLevel(logging.INFO)
                format = logging.Formatter(
                    '%(asctime)s >> %(name)s >> %(levelname)s >> %(message)s')

                fileHandler = logging.FileHandler(self.loggerFileName)
                fileHandler.setLevel(logging.INFO)
                fileHandler.setFormatter(format)

                consoleHandler = logging.StreamHandler()
                consoleHandler.setLevel(logging.INFO)
                consoleHandler.setFormatter(format)

                self.logger.addHandler(fileHandler)
                self.logger.addHandler(consoleHandler)
                
            # Initiate Driver         
            prefs = {"download.default_directory": downloadPath, 'download.directory_upgrade': True}
            chrome_option = webdriver.ChromeOptions()
            chrome_option.add_experimental_option("prefs", prefs)
            # chrome_options.add_argument('--disable-application-cache')
            self.driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_option)
            driver_wait = WebDriverWait(self.driver, 30, poll_frequency=1, ignored_exceptions=[
                               ElementNotVisibleException, ElementNotSelectableException])
            self.logger.info(f"Web Driver Initiated Successfully")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")

    def test_close_browser(self):
        self.driver.quit()

    def zipdir(self):
        try:
            currentDate = date.today()
            zipFileName = f"Recon Reports - {currentDate.strftime('%m-%d-%Y')}.zip"
            zipf = zipfile.ZipFile(zipFileName, 'w', zipfile.ZIP_DEFLATED)
            pattern = "*.csv"
            for root, dirs, files in os.walk(reconReportDownloadPath):
                for file in files:
                    if fnmatch.fnmatch(file, pattern):
                        os.chdir(reconReportDownloadPath)
                        self.logger.info(f"Adding {file} to zip file {zipFileName}")
                        zipf.write(file)
            zipf.close()
            return zipFileName
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")

    def removeFiles(self, downloadFolder):
        try:
            filelist = [f for f in os.listdir(downloadFolder) if f.endswith(".csv") or f.endswith(".zip")]
            for f in filelist:
                os.remove(os.path.join(downloadFolder, f))
                self.logger.info(f"Removing file {f}")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")

    def renameFile(self, sourceFilePath, destFileName):
        try:
            listOfFiles = os.listdir(sourceFilePath)
            pattern = "*.csv"
            for entry in listOfFiles:
                if fnmatch.fnmatch(entry, pattern):
                    os.chdir(sourceFilePath)
                    self.logger.info(f"Renaming file from {entry} to {destFileName}")
                    os.rename(entry, destFileName)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")

    def moveFile(self, sourceFilePath, destFilePath, destFileName):
        try:
            for filename in os.listdir(sourceFilePath):
                if str(filename) == destFileName:
                    shutil.move(filename, destFilePath)
                    self.logger.info(f"Moving file {filename} to {destFilePath}")

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")
            self.logger.exception(getattr(e, 'message', repr(e)))
            self.logger.exception(getattr(e, 'message', str(e)))

    def copyFile(self, sourceFilePath, destFilePath, destFileName):
        try:
            for filename in os.listdir(sourceFilePath):
                if str(filename) == destFileName:                    
                    shutil.copy(filename, destFilePath)
                    self.logger.info(f"Copying file {filename} to {destFilePath}")

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")

    def test_open_url(self):
        try:
            wait = WebDriverWait(self.driver, 720)
            if os.getenv("LOGIN_BY") == "SALESFORCE":
                if os.getenv("SANDBOX") == True:
                    self.driver.get(os.getenv("SANDBOX_URL"))
                    self.logger.info(f"Opening URL {os.getenv('SANDBOX_URL')}")
                else:
                    self.driver.get(os.getenv("PRODUCTION_URL"))
                    self.logger.info(f"Opening URL {os.getenv('PRODUCTION_URL')}")
                wait.until(EC.visibility_of_element_located((By.ID, "Login")))
            elif os.getenv("LOGIN_BY") == "OKTA":
                self.driver.get(os.getenv("OKTA_URL"))
                self.logger.info(f"Opening URL {os.getenv('OKTA_URL')}")
                wait.until(EC.visibility_of_element_located((By.ID, "okta-signin-username")))
            self.driver.maximize_window()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")

    def test_loginSalesForce(self):
        try:            
            wait = WebDriverWait(self.driver, 720)
            driver = self.driver

            if os.getenv("LOGIN_BY") == "SALESFORCE":
                txtBoxUser = driver.find_element_by_id("username")
                txtBoxpwd = driver.find_element_by_id("password")
                btnLogin = driver.find_element_by_id("Login")

                if os.getenv("SANDBOX") == True:
                    txtBoxUser.send_keys(os.getenv("SandBoxUserID"))
                    txtBoxpwd.send_keys(os.getenv("SandBoxUserPassword"))
                else:
                    txtBoxUser.send_keys(os.getenv("ProdUserID"))
                    txtBoxpwd.send_keys(os.getenv("ProdUserPassword"))
                btnLogin.click()
                self.logger.info(f"User {os.getenv('username')} logged in by {os.getenv('LOGIN_BY')}")
                wait.until(EC.visibility_of_element_located((By.XPATH, "//img[@title='All Tabs']")))
            
            elif os.getenv("LOGIN_BY") == "OKTA":
                txtBoxLogin = driver.find_element_by_id("okta-signin-username")
                btnNext = driver.find_element_by_id("okta-signin-submit")

                txtBoxLogin.send_keys(os.getenv("OKTA_ID"))
                btnNext.click()

                wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Verify']")))

                txtBoxPassword = driver.find_element_by_xpath("//input[@name='password']")
                txtBoxPassword.send_keys(os.getenv("OKTA_Password"))
                btnSignin = driver.find_element_by_xpath("//input[@value='Verify']")
                btnSignin.click()
                self.logger.info(f"User {os.getenv('OKTA_ID')} logged in by {os.getenv('LOGIN_BY')}")

                # wait.until(EC.visibility_of_element_located((By.XPATH, "//a[img[@alt='Graphic Link Salesforce']]")))
                wait.until(EC.visibility_of_element_located((By.XPATH, "//h1[text()='Salesforce']")))

                imgSalesforce = driver.find_element_by_xpath("//h1[text()='Salesforce']")

                imgSalesforce.click()
                main_window = driver.current_window_handle
                sleep(10)
                driver.switch_to.window(driver.window_handles[1])
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")

    def test_ExportReports(self):
        try:

            wait = WebDriverWait(self.driver, 720)
            driver = self.driver
            # driver.switch_to.window(driver.window_handles[-1])

            allTabs = driver.find_element_by_xpath("//img[@title='All Tabs']")
            allTabs.click()
            wait.until(EC.visibility_of_element_located((By.LINK_TEXT, "Recon Reports")))

            reconReportsLink = driver.find_element_by_xpath("//img[@title='Recon Reports']")
            reconReportsLink.click()
            self.logger.info(f"Opened the Tab Recon Reports")
            wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Generate Report']")))

            self.removeFiles(downloadPath)
            self.removeFiles(reconReportDownloadPath)
            wait = WebDriverWait(self.driver, 180, 2)
            driver = self.driver
            datem = datetime.today().strftime("%Y")
            recordCount = 0

            currentURL = self.driver.current_url

            fiscalYears = [str(int(datem)), str(int(datem) + 1), str(int(datem) + 2), str(int(datem) + 3)]

            productLines = ["Goal team have multiple Specialty", "Opportunity team have multiple Specialty",
                            "Opportunity team have multiple core Primary", "Goal team have multiple core Primary",
                            "Opportunity with no Core team", "Goal with no Core team", "Instore Oppty with Insert Date",
                            "Teams not appropriate as per Territory-Goal",
                            "Teams not appropriate as per Territory-Oppty"]

            for productLine in productLines:
                self.logger.info(f"Searching for {productLine}")

                self.recordCountMap[productLine] = 0
                wait.until(EC.invisibility_of_element_located((By.XPATH, "//img[src='/img/loading32.gif']")))
                for fiscalYear in fiscalYears:
                    self.logger.info(f"Searching Records For Fiscal Year {fiscalYear}")
                    wait.until(EC.invisibility_of_element_located((By.XPATH, "//img[src='/img/loading32.gif']")))

                    fiscalYearDropDown = Select(
                        driver.find_element_by_xpath("//div[text()='Fiscal Year:']/following-sibling::div/select"))
                    fiscalYearDropDown.select_by_value(fiscalYear)

                    productLineDropDown = Select(driver.find_element_by_xpath(
                        "//div[text()='Select Recon report:']/following-sibling::div/select"))
                    productLineDropDown.select_by_value(productLine)

                    generateReportButton = driver.find_element_by_xpath("//input[@value='Generate Report']")
                    generateReportButton.click()

                    wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))

                    xpathRecords = "//div[@id='reportDiv']//table//tr['.dataRow'] | //div[@id='tablePrint']//table//tr"
                    recordCount = len(driver.find_elements_by_xpath(xpathRecords))
                    

                    test_words = ['Account', 'appropriate']
                    if not any(test_word in productLine for test_word in test_words):
                        #                         if not test_words in productLine:
                        self.recordCountMap[productLine] = int(self.recordCountMap[productLine]) + (
                                    recordCount - 1)
                        self.logger.info(f"Record Count For {productLine} : {int(self.recordCountMap[productLine]) + (recordCount - 1)}")
                    else:
                        self.recordCountMap[productLine] = recordCount - 1
                        self.logger.info(f"Record Count For {productLine} : {recordCount - 1}")
                    
                    
                    if (recordCount - 1) >= 1:
                        self.removeFiles(downloadPath)

                        exportToExcelLink = driver.find_element_by_xpath("//a[contains(text(),'Export To Excel')]")
                        exportToExcelLink.click()
                        sleep(10)

                        destFileName = productLine + "#" + fiscalYear + ".csv"
                        self.renameFile(downloadPath, destFileName)
                        sleep(2)
                        self.copyFile(downloadPath, reconReportDownloadPath, destFileName)
                        sleep(2)
                        wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Generate Report']")))
                    self.logger.info(self.recordCountMap)

            productLines = ["Account(Division type) not have Territory", "Account team have multiple Core Primary",
                            "Account team have multiple Specialty Primary",
                            "Account team having TR after form header closed",
                            "Teams not appropriate as per Territory-Account"]

            for productLine in productLines:
                self.logger.info(f"Searching for {productLine}")

                self.recordCountMap[productLine] = 0
                wait.until(EC.invisibility_of_element_located((By.XPATH, "//img[src='/img/loading32.gif']")))

                productLineDropDown = Select(
                    driver.find_element_by_xpath("//div[text()='Select Recon report:']/following-sibling::div/select"))
                productLineDropDown.select_by_value(productLine)

                generateReportButton = driver.find_element_by_xpath("//input[@value='Generate Report']")
                generateReportButton.click()

                wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))

                xpathRecords = "//div[@id='reportDiv']//table//tr['.dataRow'] | //div[@id='tablePrint']//table//tr"
                recordCount = len(driver.find_elements_by_xpath(xpathRecords))

                test_words = ['Account', 'appropriate']
                if not any(test_word in productLine for test_word in test_words):
                    self.recordCountMap[productLine] = int(self.recordCountMap[productLine]) + (recordCount - 1)
                    self.logger.info(f"Record Count For {productLine} : {int(self.recordCountMap[productLine]) + (recordCount - 1)}")
                elif "Teams not appropriate as per Territory-Account" in productLine:
                    self.recordCountMap[productLine] = recordCount - 2
                    self.logger.info(f"Record Count For {productLine} : {recordCount - 2}")
                else:
                    self.recordCountMap[productLine] = recordCount - 1
                    self.logger.info(f"Record Count For {productLine} : {recordCount - 1}")

                if (recordCount - 1) >= 1:
                    self.removeFiles(downloadPath)

                    exportToExcelLink = driver.find_element_by_xpath("//a[contains(text(),'Export To Excel')]")
                    exportToExcelLink.click()
                    sleep(10)

                    destFileName = productLine + ".csv"
                    self.renameFile(downloadPath, destFileName)
                    sleep(2)
                    self.copyFile(downloadPath, reconReportDownloadPath, destFileName)
                    sleep(2)
                    wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Generate Report']")))
                self.logger.info(self.recordCountMap)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.test_ExportReports()
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno} \n Error Message: {e}")

    def test_sendMail(self):
        try:

            currentDate = date.today()
            currentDate = currentDate.strftime("%m.%d.%Y")

            doc, tag, text = Doc().tagtext()

            with tag('html'):
                with tag('body'):
                    with tag('font', face="Consolas"):
                        with tag('h3', align="left"):
                            text(f"RECORN REPORTS STATUS : {currentDate}")

                    with tag('font', size="2", face="Consolas"):
                        with tag('table'):
                            with tag('tbody'):
                                with tag('tr'):
                                    with tag('td', bgcolor="#AEA4A2", align="center"):
                                        text('SR NO.')
                                    with tag('td', bgcolor="#AEA4A2", align="center"):
                                        text('REPORT NAME')
                                    with tag('td', bgcolor="#AEA4A2", align="center"):
                                        text('RECORD COUNT')
                            for i, (k, v) in enumerate(self.recordCountMap.items()):
                                with tag('tr'):
                                    with tag('td', align="center", bgcolor="#FA8072" if int(v) > 0 else "#08c96b"):
                                        text(i + 1)
                                    with tag('td', align="center", bgcolor="#FA8072" if int(v) > 0 else "#08c96b"):
                                        text(k)
                                    with tag('td', align="center", bgcolor="#FA8072" if int(v) > 0 else "#08c96b"):
                                        text(f"{v} - RECORD(S) FOUND")

                    with tag('font', face="Consolas"):
                        with tag('p'):
                            text(
                                "Added all records into below 12.25.09.01.02 Salesforce Reconciliation Results Google sheet:")
                        with tag('a',
                                 href="https://docs.google.com/spreadsheets/d/1-ucCKbtAa5nJRS_ZdBO7NZkrEpoR_tf9zpD-BZFgieM/edit#gid=313202824"):
                            text(
                                "https://docs.google.com/spreadsheets/d/1-ucCKbtAa5nJRS_ZdBO7NZkrEpoR_tf9zpD-BZFgieM/edit#gid=313202824")
                        with tag('p'):
                            text("Please let me know if there is any query.")

            tableData = doc.getvalue()
            self.logger.info(tableData)
            zipFileName = self.zipdir()
            currentDate = date.today()
            smtp_server = os.getenv("SMTPServer")
            port = os.getenv("SMTPPort")
            sender_email = os.getenv("SenderEmail")
            password = os.getenv("SenderPassword")
            toaddr = os.getenv("ToEmails")
            ccaddr = os.getenv("CcEmails")
            rcpt = ccaddr.split(",") + toaddr.split(",")
            #                 bccaddr = os.getenv("BccEmails"].split(",")
            context = ssl.create_default_context()
            #                 body = """
            #                 Hi All,
            #
            #                 Please find the attached compressed zip file may having Recon Report based on issues present in records:
            #
            #
            #
            #                 Regards,
            #
            #                 SalesForce Admin """
            #
            #                 message = MIMEMultipart()

            html = """
                <html>
                <head>
                <style>
                  table, th, td {{ border: 1px solid black; border-collapse: collapse; }}
                  th, td {{ padding: 1px; }}
                </style>
                </head>
                <body><p>Hi All,</p>
                <p>Please find below today's Recon Report:</p>
                {table}
                
                <p>Added all records into Google sheet: <a href="https://docs.google.com/spreadsheets/d/1-ucCKbtAa5nJRS_ZdBO7NZkrEpoR_tf9zpD-BZFgieM/edit#gid=313202824">Google Sheet</a> </p>
                <p>Also, added Fiscal year 2021 Opportunities Reports of US/CAD into Report Extracts Folder shared with BO Team in today's date folder :
                <a href="https://drive.google.com/drive/u/0/folders/1w0pecpKe63gyrvrkBNVr1Li3hEz8aNnL">Recon Report Extracts Google Driver</a> </p> 
                <p>Regards,</p>
                <p>NAM Admin</p>
                </body></html>
                
                """
            #                 with open(cwd + "\\" + os.getenv("CSV_FILE_NAME"]) as input_file:
            #                     reader = csv.reader(input_file)
            #                     data = list(reader)

            #                 text = text.format(table=tabulate(data, headers="firstrow", tablefmt="grid", stralign="center"))
            #                 html = html.format(table=tabulate(data, headers="firstrow", tablefmt="html", stralign="center", numalign="center"))

            message = MIMEMultipart(
                "alternative", None, [MIMEText(tableData, 'html')])

            message['Subject'] = "Recon Report: " + str(currentDate.strftime('%m-%d-%Y'))
            message['From'] = sender_email
            message['To'] = toaddr
            message['Cc'] = ccaddr
            #                 message['To'] = ", ".join(toaddr)
            #                 message['Cc'] = ", ".join(ccaddr)
            #                 message['Bcc'] = ", ".join(bccaddr)

            #                 filename = downloadPath + f"{currentDate.strftime('%m-%d-%Y')}.xls"

            # Add body to email
            #                 message.attach(MIMEText(body, "html"))

            # Open PDF file in binary mode
            with open(downloadPath + zipFileName, "rb") as attachment:
                # Add file as application/octet-stream
                # Email client can usually download this automatically as attachment
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            # Encode file in ASCII characters to send by email
            encoders.encode_base64(part)

            # Add header as key/value pair to attachment part
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {zipFileName}",
            )
            #
            # Add attachment to message and convert message to string
            message.attach(part)
            text = message.as_string()

            server = smtplib.SMTP(smtp_server, port)
            server.ehlo()  # Can be omitted
            server.starttls(context=context)  # Secure the connection
            server.ehlo()  # Can be omitted
            server.login(sender_email, password)
            server.sendmail(sender_email, rcpt, text)
            server.quit()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")

    def sendErrorMail(self, traceStr):
        try:
            currentDate = date.today()
            
            smtp_server = os.getenv("SMTPServer")
            port = os.getenv("SMTPPort")

            sender_email = os.getenv("SenderEmail")
            password = os.getenv("SenderPassword")
            toaddr = os.getenv("ErrorEmails").split(",")

            self.logger.info(f"Sending error email...")
            context = ssl.create_default_context()
            body = traceStr
            message = MIMEMultipart()

            message['Subject'] = "Error in Recon Report: " + str(currentDate)
            message['From'] = sender_email
            message['ErrorEmails'] = ", ".join(toaddr)
            #                 message['Bcc'] = ", ".join(bccaddr)

            # Add body to email
            message.attach(MIMEText(body, "plain"))
            text = message.as_string()

            server = smtplib.SMTP(smtp_server, port)
            server.ehlo()  # Can be omitted
            server.starttls(context=context)  # Secure the connection
            server.ehlo()  # Can be omitted
            server.login(sender_email, password)
            server.sendmail(sender_email, toaddr, text)
            server.quit()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            
            self.logger.exception(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
            self.sendErrorMail(f"{exc_type} == {fname} == {exc_tb.tb_lineno}")
